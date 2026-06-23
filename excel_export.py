"""
excel_export.py
===============
Genera el archivo Excel con las siguientes hojas, en este orden:
  1. Datos Diarios              : tabla + gráfico embebido + Resumen Mensual debajo
  2. PEP                        : PEA del mes + SISTEMA por día/versión
  3+ {Versión} Horario          : matriz horaria del PBNA en esa versión, sin techar
     {Versión} Horario Techado  : matriz horaria de esa versión corregida con el PEP
     (un par Horario/Horario Techado por cada versión: TX1, TX2, TXR*, TXF*)
"""

from __future__ import annotations

import io
import calendar
from typing import Optional

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.legend import Legend
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────────────────────────
#  CONSTANTES DE ESTILO
# ──────────────────────────────────────────────────────────────────

_MESES_ES_ABR = {
    1: "ene", 2: "feb", 3: "mar", 4: "abr",
    5: "may", 6: "jun", 7: "jul", 8: "ago",
    9: "sep", 10: "oct", 11: "nov", 12: "dic",
}
_MESES_ES_FULL = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

def _fmt_fecha_es(value) -> str:
    """Convierte una fecha ISO a formato dd-mmm en español (ej: 01-abr)."""
    try:
        d = pd.to_datetime(str(value))
        return f"{d.day:02d}-{_MESES_ES_ABR[d.month]}"
    except Exception:
        return str(value)

C_HEADER_BG  = "1A2744"
C_HEADER_FG  = "FFFFFF"
C_ALT_ROW    = "EEF2F8"
NUM_FMT      = '#,##0.00'


# ──────────────────────────────────────────────────────────────────
#  UTILIDADES DE ESTILO
# ──────────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _thick_border(color: str = "CC0000") -> Border:
    side = Side(style="thick", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_style(ws, row: int, n_cols: int) -> None:
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font      = Font(name="Calibri", bold=True, color=C_HEADER_FG, size=10)
        cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()


def _num_cell(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if value is not None and not (isinstance(value, float) and np.isnan(value)):
        cell.number_format = NUM_FMT
    cell.alignment = Alignment(horizontal="right")
    cell.border    = _thin_border()


def _text_cell(ws, row: int, col: int, value, bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Calibri", bold=bold, size=10)
    cell.alignment = Alignment(horizontal="left")
    cell.border    = _thin_border()


def _apply_color_scale(ws, start_row: int, col: int, end_row: int) -> None:
    """
    Formato condicional suave: verde = precio bajo, rojo = precio alto
    (mismos tonos pastel que el reporte HTML).
    """
    col_letter = get_column_letter(col)
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    rule = ColorScaleRule(
        start_type="min",        start_color="63BE7B",   # verde suave  → bajo
        mid_type="percentile",   mid_value=50,
                                 mid_color="FFEB84",     # amarillo suave → medio
        end_type="max",          end_color="F8696B",     # rojo suave   → alto
    )
    ws.conditional_formatting.add(cell_range, rule)


def _add_series(chart: LineChart, ws_data, col: int,
                min_row: int, max_row: int,
                title: str, color: str,
                dash: str = "solid", width: int = 20000) -> None:
    """Añade una serie a un LineChart con título, color y marcadores circulares."""
    ref = Reference(ws_data, min_col=col, min_row=min_row, max_row=max_row)
    s = Series(ref)
    s.title = SeriesLabel(v=title)
    s.graphicalProperties.line.solidFill = color
    s.graphicalProperties.line.width     = width
    if dash != "solid":
        s.graphicalProperties.line.dashStyle = dash
    s.smooth = False
    s.marker.symbol = "circle"
    s.marker.size   = 4
    s.marker.graphicalProperties.solidFill          = color
    s.marker.graphicalProperties.line.solidFill     = color
    chart.series.append(s)


# ──────────────────────────────────────────────────────────────────
#  HOJA 1: DATOS DIARIOS
# ──────────────────────────────────────────────────────────────────

def _build_daily_columns(df: pd.DataFrame) -> list[tuple[str, str]]:
    """
    Construye la lista de (header_display, col_en_df) de forma dinámica,
    incluyendo TXR y TXF solo si están presentes y tienen datos.
    """
    has_txr = "TXR" in df.columns and df["TXR"].notna().any()
    has_txf = "TXF" in df.columns and df["TXF"].notna().any()

    cols = [
        ("Fecha",                "Fecha"),
        ("Predespacho",          "Predespacho"),
        ("TX1",                  "TX1"),
        ("TX2",                  "TX2"),
    ]
    if has_txr:
        cols.append(("TXR", "TXR"))
    if has_txf:
        cols.append(("TXF", "TXF"))
    cols += [
        ("Mejor Versión",        "Mejor_Version"),
        ("PB Diario",            "PB_Diario"),
        ("Δ Versión",            "Diferencia_Version"),
        ("Versión Nombre",       "Mejor_Version_Nombre"),
        ("Pred PromMes",         "Predespacho_PromMes"),
        ("TX1 PromMes",          "TX1_PromMes"),
        ("TX2 PromMes",          "TX2_PromMes"),
    ]
    if has_txr:
        cols.append(("TXR PromMes", "TXR_PromMes"))
    if has_txf:
        cols.append(("TXF PromMes", "TXF_PromMes"))
    cols.append(("PB PromMes", "PB_PromMes"))
    return cols


# Columnas cuyo valor se escribe como número
_NUMERIC_SRC = {
    "Predespacho", "TX1", "TX2", "TXR", "TXF",
    "Mejor_Version", "PB_Diario", "Diferencia_Version",
    "Predespacho_PromMes", "TX1_PromMes", "TX2_PromMes",
    "TXR_PromMes", "TXF_PromMes", "PB_PromMes",
}

# Columna fuente de precio → columna booleana que indica techado por PEP
_CAPPED_FLAG_SRC = {
    "TX1":           "TX1_Techado",
    "TX2":           "TX2_Techado",
    "TXR":           "TXR_Techado",
    "TXF":           "TXF_Techado",
    "Mejor_Version":  "Mejor_Version_Techado",
    "PB_Diario":      "Mejor_Version_Techado",
}
_PEP_COMMENT_TEXT = (
    "Precio techado por PEP (Precio de Escasez Ponderado): al menos una hora "
    "de este día superó el PEA y fue limitada al valor SISTEMA."
)


def _write_daily_sheet(ws, df: pd.DataFrame) -> dict[str, int]:
    """
    Escribe la hoja Datos Diarios y retorna un dict {src_col: col_index_excel}
    para que la hoja de gráficos pueda referenciar columnas por nombre.
    """
    ws.title = "Datos Diarios"

    col_defs = _build_daily_columns(df)
    headers  = [h for h, _ in col_defs]
    src_cols = [s for _, s in col_defs]

    # índice src_col → número de columna Excel (1-based)
    col_idx_map: dict[str, int] = {s: i + 1 for i, s in enumerate(src_cols)}

    # ── Encabezados ────────────────────────────────────────────────
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _header_style(ws, 1, len(headers))

    # ── Datos ──────────────────────────────────────────────────────
    data_start = 2
    has_capped = False
    for row_idx, (_, row) in enumerate(df.iterrows(), start=data_start):
        fill = PatternFill("solid", fgColor=C_ALT_ROW) if row_idx % 2 == 0 else None

        for col_idx, (h, src) in enumerate(col_defs, start=1):
            value = row.get(src)

            if src == "Fecha":
                try:
                    date_val = pd.to_datetime(str(value)).date()
                    cell = ws.cell(row=row_idx, column=col_idx, value=date_val)
                    cell.number_format = "dd-mmm"
                except Exception:
                    cell = ws.cell(row=row_idx, column=col_idx, value=_fmt_fecha_es(value))
                cell.alignment = Alignment(horizontal="center")
                cell.border    = _thin_border()
            elif src in _NUMERIC_SRC:
                raw = None if (value is None or (isinstance(value, float) and np.isnan(value))) else float(value)
                _num_cell(ws, row_idx, col_idx, raw)
                flag_col = _CAPPED_FLAG_SRC.get(src)
                if flag_col:
                    flag_val = row.get(flag_col)
                    if pd.notna(flag_val) and bool(flag_val):
                        has_capped = True
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = _thick_border()
                        cell.comment = Comment(_PEP_COMMENT_TEXT, "XM Precios")
            else:
                _text_cell(ws, row_idx, col_idx, str(value) if value else "")

            if fill:
                ws.cell(row=row_idx, column=col_idx).fill = fill

    data_end = data_start + len(df) - 1

    if has_capped:
        note_row = data_end + 1
        note_cell = ws.cell(
            row=note_row, column=1,
            value="🔻 Borde grueso rojo (con comentario): precio techado por el PEP en al menos una hora del día.",
        )
        note_cell.font = Font(name="Calibri", italic=True, size=9, color="888888")

    # ── Formato condicional en columnas de precio ───────────────────
    price_src = ["Predespacho", "TX1", "TX2", "TXR", "TXF", "Mejor_Version", "PB_Diario", "PB_PromMes"]
    for src in price_src:
        if src in col_idx_map:
            _apply_color_scale(ws, data_start, col_idx_map[src], data_end)

    # ── Anchos de columna ──────────────────────────────────────────
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            12 if h == "Fecha" else
            18 if "PromMes" in h else
            14
        )

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{data_end}"

    return col_idx_map, data_start, data_end


# ──────────────────────────────────────────────────────────────────
#  BLOQUE: RESUMEN MENSUAL (debajo de la gráfica, en Datos Diarios)
# ──────────────────────────────────────────────────────────────────

def _write_summary_block(ws, summary: dict, year: int, month: int, start_row: int) -> int:
    """
    Escribe el bloque de Resumen Mensual a partir de start_row, en la misma
    hoja de Datos Diarios (ya no es una hoja independiente). Retorna la
    última fila utilizada.
    """
    month_name = calendar.month_name[month]

    # Título
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
    cell = ws.cell(row=start_row, column=1)
    cell.value     = f"Resumen Mensual — {month_name} {year}"
    cell.font      = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 28

    header_row = start_row + 1
    for col_idx, h in enumerate(["Métrica", "Valor ($/kWh)", "Notas"], start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _header_style(ws, header_row, 3)

    notes = {
        "Promedio Predespacho":   "Promedio archivos IMAR (Costo Marginal / 1000)",
        "Promedio TX1":           "Primera liquidación provisional (TRSD .tx1 — PBNA)",
        "Promedio TX2":           "Segunda liquidación provisional (TRSD .tx2 — PBNA)",
        "Promedio TXR":           "Reliquidación (TRSD .txr) — disponible día 5 mes siguiente",
        "Promedio TXF":           "Liquidación final (TRSD .txf) — disponible día 10 mes siguiente",
        "Promedio Mejor Versión": "Mejor versión disponible: TXF > TXR > TX2 > TX1 > Predespacho",
        "Promedio PB Mes":        "Precio de Bolsa mensual (igual a Mejor Versión)",
    }

    fills = [
        PatternFill("solid", fgColor="EEF2F8"),
        PatternFill("solid", fgColor="FFFFFF"),
    ]

    row_idx = header_row
    for key, val in summary.items():
        row_idx += 1
        alt = fills[row_idx % 2]
        for c in range(1, 4):
            ws.cell(row=row_idx, column=c).fill = alt
        _text_cell(ws, row_idx, 1, key, bold=True)
        _num_cell(ws, row_idx, 2, val if val is not None else None)
        _text_cell(ws, row_idx, 3, notes.get(key, ""))

    return row_idx


# ──────────────────────────────────────────────────────────────────
#  HOJA 3: GRÁFICOS
# ──────────────────────────────────────────────────────────────────

def _write_chart_in_sheet(
    ws,
    col_idx_map: dict[str, int],
    data_start: int,
    data_end: int,
    year: int,
    month: int,
    y_min: Optional[float] = None,
) -> None:

    all_series = [
        ("Predespacho",   "Predespacho",      "FFB300", "solid", 15000),
        ("TX1",           "TX1",              "FF6600", "solid", 15000),
        ("TX2",           "TX2",              "00AA2E", "solid", 15000),
        ("TXR",           "TXR",              "5BA300", "solid", 15000),
        ("TXF",           "TXF",              "B8860B", "solid", 15000),
        ("Mejor_Version", "Mejor Versión",    "CC0000", "solid", 20000),
        ("PB_Diario",     "PB Diario",        "2E75B6", "solid", 15000),
        ("PB_PromMes",    "Promedio Mensual", "FF4500", "dash",  25000),
    ]

    chart = LineChart()
    chart.style        = 2          # fondo blanco, limpio
    chart.grouping     = "standard"
    chart.smooth       = False
    chart.height       = 22
    chart.width        = 38

    # ── Título ────────────────────────────────────────────────────
    chart.title = f"Precio de Bolsa {_MESES_ES_FULL[month]} {year}  —  Versiones y Promedio Mensual"

    # ── Ejes ──────────────────────────────────────────────────────
    chart.y_axis.title      = "Precio ($/kWh)"
    chart.y_axis.numFmt     = '#,##0.00'
    chart.y_axis.tickLblPos = "nextTo"
    chart.y_axis.delete     = False
    if y_min is not None:
        chart.y_axis.scaling.min = float(y_min)

    chart.x_axis.title       = "Fecha"
    chart.x_axis.numFmt      = "dd-mmm"
    chart.x_axis.tickLblPos  = "nextTo"
    chart.x_axis.delete      = False
    chart.x_axis.tickLblSkip = 1

    # ── Área de trazado con márgenes explícitos ───────────────────
    # Margen izquierdo amplio → título eje Y no se solapa con las cifras.
    # Altura reducida        → título eje X queda debajo de las fechas.
    try:
        from openpyxl.chart.layout import Layout, ManualLayout
        chart.plot_area.layout = Layout(
            manualLayout=ManualLayout(
                x=0.15,   # desplazamiento izq. (espacio para título eje Y)
                y=0.04,   # margen superior
                w=0.78,   # ancho del área de trazado
                h=0.72,   # alto (deja ~24 % abajo para etiquetas + título eje X)
            )
        )
    except Exception:
        pass

    # ── Grid suave: gris muy claro ────────────────────────────────
    try:
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        _grid = GraphicalProperties(ln=LineProperties(solidFill="EBEBEB"))
        chart.y_axis.majorGridlines.spPr = _grid
    except Exception:
        pass

    # ── Leyenda fuera del área, esquina superior derecha ──────────
    lgnd = Legend()
    lgnd.position = "r"
    lgnd.overlay  = False
    chart.legend  = lgnd

    # ── Series ────────────────────────────────────────────────────
    for src, label, color, dash, width in all_series:
        if src in col_idx_map:
            _add_series(
                chart, ws,
                col=col_idx_map[src],
                min_row=data_start, max_row=data_end,
                title=label, color=color,
                dash=dash, width=width,
            )

    # ── Categorías del eje X (debe ir DESPUÉS de agregar las series) ──
    fecha_col = col_idx_map.get("Fecha", 1)
    cats = Reference(ws, min_col=fecha_col, min_row=data_start, max_row=data_end)
    chart.set_categories(cats)

    ws.add_chart(chart, f"A{data_end + 3}")


# Filas que ocupa aprox. una gráfica de chart.height=22 (cm), a ~0.53 cm/fila,
# más un margen — usado para ubicar el Resumen Mensual debajo de la gráfica.
_CHART_HEIGHT_ROWS = 44

NUM_HOURS = 24


# ──────────────────────────────────────────────────────────────────
#  HOJAS: MATRICES HORARIAS TX1 (PBNA cruda / techada)
# ──────────────────────────────────────────────────────────────────

_PEP_FILL = PatternFill("solid", fgColor="FFD9A0")  # naranja claro — hora techada


def _write_hourly_matrix_sheet(
    ws,
    df: pd.DataFrame,
    hours_col: str,
    compare_col: Optional[str] = None,
) -> None:
    """
    Escribe una matriz horaria: Fecha | H1..H24 | Promedio, una fila por día.

    `hours_col` es la columna del df con listas de 24 valores (PBNA TX1).
    Si `compare_col` se indica, las horas cuyo valor difiera respecto a esa
    otra columna (es decir, horas techadas por el PEP) se resaltan con
    relleno naranja y un comentario explicativo.
    """
    headers = ["Fecha"] + [f"H{h}" for h in range(1, NUM_HOURS + 1)] + ["Promedio"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    _header_style(ws, 1, len(headers))

    data_start = 2
    for row_idx, (_, row) in enumerate(df.iterrows(), start=data_start):
        alt_fill = PatternFill("solid", fgColor=C_ALT_ROW) if row_idx % 2 == 0 else None

        date_val = row.get("Fecha")
        try:
            d = pd.to_datetime(str(date_val)).date()
            cell = ws.cell(row=row_idx, column=1, value=d)
            cell.number_format = "dd-mmm"
        except Exception:
            cell = ws.cell(row=row_idx, column=1, value=str(date_val))
        cell.alignment = Alignment(horizontal="center")
        cell.border = _thin_border()
        if alt_fill:
            cell.fill = alt_fill

        hours = row.get(hours_col)
        hours = hours if isinstance(hours, list) else [None] * NUM_HOURS
        compare_hours = None
        if compare_col:
            ch = row.get(compare_col)
            compare_hours = ch if isinstance(ch, list) else None

        for h_idx in range(NUM_HOURS):
            col_idx = 2 + h_idx
            val = hours[h_idx] if h_idx < len(hours) else None
            raw_val = None if (val is None or (isinstance(val, float) and np.isnan(val))) else float(val)
            _num_cell(ws, row_idx, col_idx, raw_val)
            cell = ws.cell(row=row_idx, column=col_idx)

            is_capped = False
            if compare_hours is not None and raw_val is not None and h_idx < len(compare_hours):
                other = compare_hours[h_idx]
                if other is not None and not (isinstance(other, float) and np.isnan(other)):
                    is_capped = abs(float(other) - raw_val) > 1e-6

            if is_capped:
                cell.fill = _PEP_FILL
                cell.comment = Comment(_PEP_COMMENT_TEXT, "XM Precios")
            elif alt_fill:
                cell.fill = alt_fill

        nums = [v for v in hours if v is not None and not (isinstance(v, float) and np.isnan(v))]
        avg = float(np.mean(nums)) if nums else None
        avg_col = len(headers)
        _num_cell(ws, row_idx, avg_col, avg)
        if alt_fill:
            ws.cell(row=row_idx, column=avg_col).fill = alt_fill

    data_end = data_start + len(df) - 1

    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{data_end}"


# ──────────────────────────────────────────────────────────────────
#  HOJA: PEP (PEA + SISTEMA por día/versión)
# ──────────────────────────────────────────────────────────────────

def _write_pep_sheet(ws, df: pd.DataFrame) -> None:
    """
    Escribe el PEA (umbral mensual de activación) y, debajo, una tabla con
    el valor de SISTEMA por día para cada versión disponible (TX1, TX2,
    TXR, TXF).
    """
    ws.cell(row=1, column=1, value="PEA (Precio de Escasez de Activación)")
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=11)
    pea_val = None
    if "PEA" in df.columns and df["PEA"].notna().any():
        pea_val = float(df["PEA"].dropna().iloc[0])
    _num_cell(ws, 1, 2, pea_val)
    ws.cell(row=1, column=2).font = Font(name="Calibri", bold=True, size=11)

    ws.cell(row=2, column=1, value="PEP (Precio de Escasez Ponderado)")
    ws.cell(row=2, column=1).font = Font(name="Calibri", bold=True, size=11)

    has_txr = "TXR_Sistema" in df.columns and df["TXR_Sistema"].notna().any()
    has_txf = "TXF_Sistema" in df.columns and df["TXF_Sistema"].notna().any()

    cols = [("Fecha", "Fecha"), ("TX1", "TX1_Sistema"), ("TX2", "TX2_Sistema")]
    if has_txr:
        cols.append(("TXR", "TXR_Sistema"))
    if has_txf:
        cols.append(("TXF", "TXF_Sistema"))

    header_row = 3
    headers = [h for h, _ in cols]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _header_style(ws, header_row, len(headers))

    data_start = header_row + 1
    for row_idx, (_, row) in enumerate(df.iterrows(), start=data_start):
        fill = PatternFill("solid", fgColor=C_ALT_ROW) if row_idx % 2 == 0 else None
        for col_idx, (h, src) in enumerate(cols, start=1):
            value = row.get(src)
            if src == "Fecha":
                try:
                    d = pd.to_datetime(str(value)).date()
                    cell = ws.cell(row=row_idx, column=col_idx, value=d)
                    cell.number_format = "dd-mmm"
                except Exception:
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.alignment = Alignment(horizontal="center")
                cell.border = _thin_border()
            else:
                raw = None if (value is None or (isinstance(value, float) and np.isnan(value))) else float(value)
                _num_cell(ws, row_idx, col_idx, raw)
            if fill:
                ws.cell(row=row_idx, column=col_idx).fill = fill

    data_end = data_start + len(df) - 1

    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.freeze_panes = f"B{data_start}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{data_end}"


# ──────────────────────────────────────────────────────────────────
#  FUNCIÓN PÚBLICA PRINCIPAL
# ──────────────────────────────────────────────────────────────────

def generate_excel(
    df: pd.DataFrame,
    summary: dict,
    year: int,
    month: int,
) -> bytes:
    """
    Genera el workbook Excel completo y retorna los bytes para
    descargar desde Streamlit (st.download_button).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    has_txr = "TXR_Horas_Crudas" in df.columns and df["TXR_Horas_Crudas"].notna().any()
    has_txf = "TXF_Horas_Crudas" in df.columns and df["TXF_Horas_Crudas"].notna().any()

    # ── Creación de hojas en el orden final deseado ─────────────────
    ws1    = wb.create_sheet("Datos Diarios")
    ws_pep = wb.create_sheet("PEP")

    ws_tx1, ws_tx1_te = wb.create_sheet("TX1 Horario"), wb.create_sheet("TX1 Horario Techado")
    ws_tx2, ws_tx2_te = wb.create_sheet("TX2 Horario"), wb.create_sheet("TX2 Horario Techado")
    ws_txr = ws_txr_te = ws_txf = ws_txf_te = None
    if has_txr:
        ws_txr, ws_txr_te = wb.create_sheet("TXR Horario"), wb.create_sheet("TXR Horario Techado")
    if has_txf:
        ws_txf, ws_txf_te = wb.create_sheet("TXF Horario"), wb.create_sheet("TXF Horario Techado")

    # Límite inferior del eje Y: múltiplo de 50 por debajo del precio mínimo
    price_cols_for_min = [c for c in ["Predespacho", "TX1", "TX2", "TXR", "TXF", "Mejor_Version"] if c in df.columns]
    y_min: Optional[float] = None
    if price_cols_for_min:
        min_val = df[price_cols_for_min].min(skipna=True).min()
        if pd.notna(min_val):
            y_min = float(int(min_val // 50) * 50)

    col_idx_map, data_start, data_end = _write_daily_sheet(ws1, df)
    _write_chart_in_sheet(ws1, col_idx_map, data_start, data_end, year, month, y_min=y_min)

    chart_anchor_row = data_end + 3
    summary_start_row = chart_anchor_row + _CHART_HEIGHT_ROWS + 2
    _write_summary_block(ws1, summary, year, month, start_row=summary_start_row)

    _write_pep_sheet(ws_pep, df)

    _write_hourly_matrix_sheet(ws_tx1, df, "TX1_Horas_Crudas")
    _write_hourly_matrix_sheet(ws_tx1_te, df, "TX1_Horas_Techadas", compare_col="TX1_Horas_Crudas")
    _write_hourly_matrix_sheet(ws_tx2, df, "TX2_Horas_Crudas")
    _write_hourly_matrix_sheet(ws_tx2_te, df, "TX2_Horas_Techadas", compare_col="TX2_Horas_Crudas")
    if has_txr:
        _write_hourly_matrix_sheet(ws_txr, df, "TXR_Horas_Crudas")
        _write_hourly_matrix_sheet(ws_txr_te, df, "TXR_Horas_Techadas", compare_col="TXR_Horas_Crudas")
    if has_txf:
        _write_hourly_matrix_sheet(ws_txf, df, "TXF_Horas_Crudas")
        _write_hourly_matrix_sheet(ws_txf_te, df, "TXF_Horas_Techadas", compare_col="TXF_Horas_Crudas")

    wb.properties.title   = f"Precio de Bolsa XM — {calendar.month_name[month]} {year}"
    wb.properties.creator = "XM Precios App"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
